import asyncio, logging, os
from collections import defaultdict
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError
from dotenv import load_dotenv
from aiogram import Bot, Dispatcher, F, Router
from aiogram.exceptions import TelegramAPIError
from aiogram.filters import Command, CommandObject
from aiogram.types import Message, CallbackQuery, InlineKeyboardMarkup, InlineKeyboardButton, BufferedInputFile, BotCommand, ReplyKeyboardMarkup, KeyboardButton
from openpyxl import Workbook
from openpyxl.styles import Font
from database import Database
from parsing import parse_expense, parse_day, parse_period

load_dotenv()
TOKEN=os.getenv('BOT_TOKEN','')
ENCRYPTION_KEY=os.getenv('ENCRYPTION_KEY','')
ALLOWED={int(x) for x in os.getenv('ALLOWED_USER_IDS','').split(',') if x.strip().isdigit()}
db=Database(os.getenv('DATABASE_PATH','data/expenses.sqlite3'), os.getenv('DEFAULT_TIMEZONE','Asia/Yekaterinburg'), ENCRYPTION_KEY)
router=Router(); pending={}; awaiting_category=set(); awaiting_date={}; awaiting_amount={}; awaiting_currency=set(); awaiting_keyword={}; awaiting_keyword_rename={}; awaiting_report_period=set(); awaiting_expense_day=set(); pending_report_range={}; day_override={}

MENU_COMMANDS = [
    BotCommand(command='start', description='👋 Начать работу с ботом'),
    BotCommand(command='help', description='🧭 Все команды и примеры'),
    BotCommand(command='categories', description='🏷️ Мои категории'),
    BotCommand(command='day', description='📅 Траты за выбранный день'),
    BotCommand(command='report', description='📊 Отчёт и Excel-файл'),
    BotCommand(command='settings', description='⚙️ Валюта и настройки'),
    BotCommand(command='newcategory', description='➕ Добавить категорию'),
    BotCommand(command='currency', description='💱 Изменить валюту'),
]

MAIN_MENU = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text='💸 Добавить трату'), KeyboardButton(text='📊 Отчёт')],
        [KeyboardButton(text='✏️ Редактировать траты')],
        [KeyboardButton(text='🏷️ Категории'), KeyboardButton(text='⚙️ Настройки')],
        [KeyboardButton(text='🧭 Помощь')],
    ],
    resize_keyboard=True,
    is_persistent=True,
    input_field_placeholder='Например: кофе 250',
)

def allowed(user_id): return user_id in ALLOWED
async def guard(message: Message):
    if not allowed(message.from_user.id): await message.answer('Доступ запрещён.'); return False
    await db.ensure_user(message.from_user.id); return True
def reset_input_state(user_id, keep_pending=False):
    awaiting_category.discard(user_id); awaiting_currency.discard(user_id); awaiting_report_period.discard(user_id); awaiting_expense_day.discard(user_id)
    awaiting_date.pop(user_id,None); awaiting_amount.pop(user_id,None); awaiting_keyword.pop(user_id,None); awaiting_keyword_rename.pop(user_id,None); day_override.pop(user_id,None)
    pending_report_range.pop(user_id,None)
    if not keep_pending: pending.pop(user_id,None)
def categories_kb(rows, prefix='choose'):
    buttons=[InlineKeyboardButton(text=r[1], callback_data=f'{prefix}:{r[0]}') for r in rows]
    return InlineKeyboardMarkup(inline_keyboard=[buttons[i:i+2] for i in range(0,len(buttons),2)]+[[InlineKeyboardButton(text='➕ Новая категория',callback_data='newcat')]])
def saved_expense_kb(expense_id):
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text='📅 Редактировать дату',callback_data=f'editdate:{expense_id}')],
        [InlineKeyboardButton(text='📂 Изменить категорию',callback_data=f'expensecat:{expense_id}')],
    ])
def currency_kb():
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text='₽ Рубли',callback_data='currency:RUB'), InlineKeyboardButton(text='$ Доллары',callback_data='currency:USD')],
        [InlineKeyboardButton(text='₸ Тенге',callback_data='currency:KZT')],
        [InlineKeyboardButton(text='✍️ Ввести вручную',callback_data='currency:custom')],
    ])
def settings_kb():
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text='💱 Изменить валюту',callback_data='settings:currency')],
    ])
def report_period_kb():
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text='За день',callback_data='report:today'), InlineKeyboardButton(text='За неделю',callback_data='report:week')],
        [InlineKeyboardButton(text='За месяц',callback_data='report:month'), InlineKeyboardButton(text='За год',callback_data='report:year')],
        [InlineKeyboardButton(text='🗓 Выбрать период',callback_data='report:custom')],
    ])
def report_category_kb(categories):
    buttons=[InlineKeyboardButton(text=category[1],callback_data=f'reportcat:{category[0]}') for category in categories]
    return InlineKeyboardMarkup(inline_keyboard=[[InlineKeyboardButton(text='Все категории',callback_data='reportcat:all')]]+[buttons[i:i+2] for i in range(0,len(buttons),2)])
def edit_expenses_period_kb():
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text='За сегодня',callback_data='editexpenses:today'), InlineKeyboardButton(text='За вчера',callback_data='editexpenses:yesterday')],
        [InlineKeyboardButton(text='🗓 За выбранную дату',callback_data='editexpenses:custom')],
    ])
def expense_categories_kb(categories,expense_id):
    buttons=[InlineKeyboardButton(text=category[1],callback_data=f'expensecatset:{expense_id}:{category[0]}') for category in categories]
    return InlineKeyboardMarkup(inline_keyboard=[buttons[i:i+2] for i in range(0,len(buttons),2)])
def categories_manage_kb(rows):
    buttons=[InlineKeyboardButton(text=r[1],callback_data=f'catview:{r[0]}') for r in rows]
    return InlineKeyboardMarkup(inline_keyboard=[buttons[i:i+2] for i in range(0,len(buttons),2)]+[[InlineKeyboardButton(text='➕ Новая категория',callback_data='catnew')]])
def category_manage_kb(category_id):
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text='➕ Ключевые слова',callback_data=f'catkwadd:{category_id}')],
        [InlineKeyboardButton(text='✏️ Изменить ключевые слова',callback_data=f'catkwlist:{category_id}')],
        [InlineKeyboardButton(text='🗑 Удалить категорию',callback_data=f'catdelete:{category_id}')],
        [InlineKeyboardButton(text='⬅️ К списку',callback_data='catlist')],
    ])
def category_delete_kb(category_id):
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text='Да, удалить',callback_data=f'catdeleteyes:{category_id}'), InlineKeyboardButton(text='Отмена',callback_data=f'catview:{category_id}')],
    ])
def category_keywords_kb(category_id,keywords):
    rows=[[InlineKeyboardButton(text=f'✏️ {keyword}',callback_data=f'catkwedit:{category_id}:{index}')]
          for index,keyword in enumerate(keywords)]
    rows.append([InlineKeyboardButton(text='⬅️ К категории',callback_data=f'catview:{category_id}')])
    return InlineKeyboardMarkup(inline_keyboard=rows)
def keyword_manage_kb(category_id,index):
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text='✏️ Изменить текст',callback_data=f'catkwrename:{category_id}:{index}')],
        [InlineKeyboardButton(text='📂 Переместить в категорию',callback_data=f'catkwmove:{category_id}:{index}')],
        [InlineKeyboardButton(text='🗑 Удалить',callback_data=f'catkwdelete:{category_id}:{index}')],
        [InlineKeyboardButton(text='⬅️ К ключевым словам',callback_data=f'catkwlist:{category_id}')],
    ])
def keyword_move_kb(source_category_id,index,categories):
    rows=[[InlineKeyboardButton(text=category[1],callback_data=f'catkwmoveto:{source_category_id}:{index}:{category[0]}')]
          for category in categories if category[0] != source_category_id]
    rows.append([InlineKeyboardButton(text='⬅️ Назад',callback_data=f'catkwedit:{source_category_id}:{index}')])
    return InlineKeyboardMarkup(inline_keyboard=rows)
def keyword_delete_kb(category_id,index):
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text='Да, удалить',callback_data=f'catkwdeleteyes:{category_id}:{index}'), InlineKeyboardButton(text='Отмена',callback_data=f'catkwedit:{category_id}:{index}')],
    ])
def day_kb(rows, expense_date):
    buttons=[]
    for r in rows:
        buttons.append([InlineKeyboardButton(text=f'✏️ {r[2][:18]} — {r[1]:g} {r[4]}',callback_data=f'editamount:{r[0]}')])
        buttons.append([InlineKeyboardButton(text='📂 Категория',callback_data=f'expensecat:{r[0]}'), InlineKeyboardButton(text='🗑 Удалить',callback_data=f'delete:{r[0]}')])
    buttons.append([InlineKeyboardButton(text='➕ Добавить трату за этот день',callback_data=f'addday:{expense_date}')])
    return InlineKeyboardMarkup(inline_keyboard=buttons)

async def reminder_loop(bot: Bot):
    """Delivers one light-hearted reminder to each user at 22:00 local time."""
    while True:
        now_utc=datetime.now(timezone.utc)
        for user_id, user_timezone in await db.reminder_users():
            try: local_now=now_utc.astimezone(ZoneInfo(user_timezone))
            except ZoneInfoNotFoundError: local_now=now_utc.astimezone(ZoneInfo('Asia/Yekaterinburg'))
            if local_now.hour == 22 and await db.claim_reminder(user_id, local_now.date().isoformat()):
                try:
                    await bot.send_message(user_id, '🧾 Псс… кошелёк просил передать: пора вспомнить, куда сегодня убежали деньги.\n\nЗапиши траты, пока они не превратились в финансовый детектив 🕵️‍♂️', reply_markup=MAIN_MENU)
                except Exception:
                    logging.exception('Could not send reminder to user %s', user_id)
        await asyncio.sleep(60)

@router.message(Command('start'))
async def start(m:Message):
    if await guard(m): reset_input_state(m.from_user.id); await m.answer('Привет! Отправьте трату: <b>Хлеб 200</b> или <b>Такси 1500 #Транспорт</b>.\n\nВсе действия — в меню внизу 👇', parse_mode='HTML', reply_markup=MAIN_MENU)
@router.message(Command('help'))
async def help_(m:Message):
    if await guard(m): reset_input_state(m.from_user.id); await m.answer('🧭 <b>Меню команд</b>\n\n🏷️ /categories — категории\n📅 /day 03.08 — траты за выбранный день\n📊 /report — выбрать период отчёта\n⚙️ /settings — валюта\n➕ /newcategory — добавить категорию\n💱 /currency — выбрать валюту\n\n<b>Быстрый ввод:</b>\n«Хлеб 200», «200 хлеб», «Хлеб 200 вчера».', parse_mode='HTML', reply_markup=MAIN_MENU)
@router.message(Command('categories'))
async def cats(m:Message):
    if not await guard(m): return
    reset_input_state(m.from_user.id); rows=await db.categories(m.from_user.id)
    await m.answer('🏷️ Выберите категорию для управления ключевыми словами или удаления:',reply_markup=categories_manage_kb(rows))
@router.message(Command('newcategory'))
async def newcat_cmd(m:Message, command:CommandObject):
    if not await guard(m): return
    reset_input_state(m.from_user.id)
    name=(command.args or '').strip()
    if not name: awaiting_category.add(m.from_user.id); return await m.answer('Введите название новой категории.')
    try: await db.add_category(m.from_user.id,name); await m.answer(f'Категория «{name}» добавлена.')
    except Exception: await m.answer('Такая категория уже есть.')
@router.message(Command('settings'))
async def settings(m:Message):
    if not await guard(m): return
    reset_input_state(m.from_user.id)
    await m.answer(f'⚙️ Настройки\n\nВалюта по умолчанию: {await db.currency(m.from_user.id)}',reply_markup=settings_kb())
@router.message(Command('currency'))
async def currency(m:Message, command:CommandObject):
    if not await guard(m): return
    reset_input_state(m.from_user.id)
    c=(command.args or '').strip()
    if not c: return await m.answer('Выберите валюту или введите свою:',reply_markup=currency_kb())
    await db.set_currency(m.from_user.id,c); await m.answer(f'Валюта изменена на {c}.')
@router.message(Command('day'))
async def day(m:Message, command:CommandObject):
    if not await guard(m): return
    reset_input_state(m.from_user.id)
    try: d=parse_day(command.args or '')
    except Exception: return await m.answer('Дата: /day 03.08 или /day 2026-08-03')
    await show_day_expenses(m,m.from_user.id,d)
async def show_day_expenses(m:Message,user_id,d):
    rows=await db.expenses(user_id,d,d)
    text=f'Траты за {d}:\n'+'\n'.join(f'{r[0]}. {r[2]} — {r[1]:g} {r[4]} ({r[5]})' for r in rows) if rows else f'За {d} трат пока нет.'
    await m.answer(text,reply_markup=day_kb(rows,d))
@router.message(Command('report'))
async def report(m:Message, command:CommandObject):
    if not await guard(m): return
    reset_input_state(m.from_user.id)
    arg=(command.args or '').lower().strip()
    if not arg: return await m.answer('📊 Выберите период отчёта:',reply_markup=report_period_kb())
    try: start,end=report_dates(arg)
    except ValueError:return await m.answer('Выберите период кнопками или укажите: /report week, /report month, /report year.')
    await ask_report_category(m,m.from_user.id,start,end)
def report_dates(arg):
    today=date.today()
    if arg in ('today','сегодня'): start=end=today
    elif arg in ('week','неделя'): start=today-timedelta(days=today.weekday()); end=today
    elif arg in ('month','месяц'): start=today.replace(day=1); end=today
    elif arg in ('year','год'): start=today.replace(month=1,day=1); end=today
    else: start,end=parse_period(arg)
    return start,end
async def ask_report_category(m:Message,user_id,start,end):
    pending_report_range[user_id]=(start,end); categories=await db.categories(user_id)
    await m.answer('📂 Выберите категорию для отчёта:',reply_markup=report_category_kb(categories))
async def send_report(m:Message,user_id,start,end,category_id=None,category_name=None):
    rows=await db.expenses(user_id,start.isoformat(),end.isoformat(),category_id)
    scope=f' по категории «{category_name}»' if category_name else ''
    if not rows:return await m.answer(f'За период {start:%d.%m.%Y}–{end:%d.%m.%Y}{scope} трат нет.')
    totals=defaultdict(float)
    for r in rows: totals[r[5]]+=r[1]
    total=sum(totals.values()); currency=rows[0][4]
    top=aggregate_top_expenses(rows,await db.categories(user_id))
    top_text='\n'.join(f'• {item[1]}{f" × {item[4]}" if item[4]>1 else ""} — {item[0]:g} {item[2]} ({item[3]})' for item in top)
    text=f'📊 Отчёт{scope}: {start:%d.%m.%Y}–{end:%d.%m.%Y}\n💰 <b>Общая сумма трат: {total:g} {currency}</b>\n\n'+'\n'.join(f'• {k}: {v:g} {currency} ({v/total:.0%})' for k,v in sorted(totals.items(),key=lambda x:-x[1]))+'\n\nТоп трат:\n'+top_text
    await m.answer(text,parse_mode='HTML'); await m.answer_document(BufferedInputFile(make_xlsx(rows,totals),'report.xlsx'),caption=f'Excel-выгрузка • Итого: {total:g} {currency}')
def aggregate_top_expenses(rows,categories,limit=5):
    keywords_by_category={category[1]:[keyword.strip() for keyword in category[2].split(',') if keyword.strip()] for category in categories}
    grouped={}; singles=[]
    for row in rows:
        matches=[keyword for keyword in keywords_by_category.get(row[5],[]) if keyword.casefold() in row[2].casefold()]
        if not matches:
            singles.append([row[1],row[2],row[4],row[5],1]); continue
        keyword=matches[0]; key=(row[5],keyword.casefold(),row[4])
        if key not in grouped: grouped[key]=[0.0,keyword,row[4],row[5],0]
        grouped[key][0]+=row[1]; grouped[key][4]+=1
    return sorted([*grouped.values(),*singles],key=lambda item:-item[0])[:limit]
def make_xlsx(rows,totals):
    wb=Workbook(); ws=wb.active; ws.title='Траты'; ws.append(['Дата','Описание','Сумма','Валюта','Категория'])
    for r in rows: ws.append([r[3],r[2],r[1],r[4],r[5]])
    total=sum(totals.values()); currency=rows[0][4]
    ws.append([]); ws.append(['ИТОГО','',total,currency,'']); ws.cell(ws.max_row,1).font=Font(bold=True); ws.cell(ws.max_row,3).font=Font(bold=True)
    ss=wb.create_sheet('Сводка'); ss.append(['Категория','Сумма','Валюта'])
    for k,v in totals.items():ss.append([k,v,currency])
    ss.append([]); ss.append(['ИТОГО',total,currency]); ss.cell(ss.max_row,1).font=Font(bold=True); ss.cell(ss.max_row,2).font=Font(bold=True)
    from io import BytesIO
    out=BytesIO(); wb.save(out); return out.getvalue()
@router.callback_query(F.data.startswith('delete:'))
async def delete(c:CallbackQuery):
    if not allowed(c.from_user.id): return
    reset_input_state(c.from_user.id)
    await db.delete_expense(c.from_user.id,int(c.data.split(':')[1])); await c.message.edit_text('Трата удалена.'); await c.answer()
@router.callback_query(F.data.startswith('editamount:'))
async def edit_amount(c: CallbackQuery):
    if not allowed(c.from_user.id): return
    reset_input_state(c.from_user.id)
    awaiting_amount[c.from_user.id]=int(c.data.split(':')[1])
    await c.message.answer('✏️ Введите новую сумму, например: 350.50')
    await c.answer()
@router.callback_query(F.data.startswith('editdate:'))
async def edit_date(c: CallbackQuery):
    if not allowed(c.from_user.id): return
    reset_input_state(c.from_user.id)
    awaiting_date[c.from_user.id]=int(c.data.split(':')[1])
    await c.message.answer('📅 Введите дату: <b>03.08</b>, <b>03.08.2026</b> или <b>2026-08-03</b>.', parse_mode='HTML')
    await c.answer()
@router.callback_query(F.data.startswith('addday:'))
async def add_day(c: CallbackQuery):
    if not allowed(c.from_user.id): return
    reset_input_state(c.from_user.id)
    day_override[c.from_user.id]=c.data.split(':',1)[1]
    await c.message.answer(f'➕ Введите трату — сохраню её за {day_override[c.from_user.id]}.')
    await c.answer()
@router.callback_query(F.data=='settings:currency')
async def settings_currency(c:CallbackQuery):
    if not allowed(c.from_user.id): return await c.answer('Доступ запрещён.',show_alert=True)
    reset_input_state(c.from_user.id)
    await c.message.edit_text('💱 Выберите валюту или введите свою:',reply_markup=currency_kb()); await c.answer()
@router.callback_query(F.data.startswith('currency:'))
async def choose_currency(c: CallbackQuery):
    if not allowed(c.from_user.id): return await c.answer('Доступ запрещён.',show_alert=True)
    await db.ensure_user(c.from_user.id)
    choice=c.data.split(':',1)[1]
    if choice=='custom':
        reset_input_state(c.from_user.id)
        awaiting_currency.add(c.from_user.id)
        await c.message.answer('✍️ Введите обозначение валюты, например: <b>€</b>, <b>EUR</b> или <b>сом</b>.',parse_mode='HTML')
        return await c.answer()
    currencies={'RUB':'₽','USD':'$','KZT':'₸'}
    selected=currencies[choice]; reset_input_state(c.from_user.id)
    await db.set_currency(c.from_user.id,selected)
    await c.message.answer(f'✅ Валюта изменена на {selected}.')
    await c.answer()
@router.callback_query(F.data.startswith('report:'))
async def choose_report_period(c:CallbackQuery):
    if not allowed(c.from_user.id): return await c.answer('Доступ запрещён.',show_alert=True)
    await db.ensure_user(c.from_user.id); reset_input_state(c.from_user.id); choice=c.data.split(':',1)[1]
    if choice=='custom':
        awaiting_report_period.add(c.from_user.id)
        await c.message.answer('🗓 Введите период в формате:\n<b>01.08.2026 - 15.08.2026</b>',parse_mode='HTML'); return await c.answer()
    try: start,end=report_dates(choice)
    except ValueError:return await c.answer('Не удалось определить период.',show_alert=True)
    await c.answer(); await ask_report_category(c.message,c.from_user.id,start,end)
@router.callback_query(F.data.startswith('reportcat:'))
async def choose_report_category(c:CallbackQuery):
    if not allowed(c.from_user.id): return await c.answer('Доступ запрещён.',show_alert=True)
    report_range=pending_report_range.get(c.from_user.id)
    if not report_range: return await c.answer('Сначала выберите период отчёта.',show_alert=True)
    choice=c.data.split(':',1)[1]; start,end=report_range
    if choice=='all':
        pending_report_range.pop(c.from_user.id,None)
        await c.answer(); return await send_report(c.message,c.from_user.id,start,end)
    category=await db.category(c.from_user.id,int(choice))
    if not category: return await c.answer('Категория не найдена.',show_alert=True)
    pending_report_range.pop(c.from_user.id,None)
    await c.answer(); await send_report(c.message,c.from_user.id,start,end,category[0],category[1])
@router.callback_query(F.data.startswith('editexpenses:'))
async def choose_expense_day(c:CallbackQuery):
    if not allowed(c.from_user.id): return await c.answer('Доступ запрещён.',show_alert=True)
    await db.ensure_user(c.from_user.id); reset_input_state(c.from_user.id); choice=c.data.split(':',1)[1]
    if choice=='custom':
        awaiting_expense_day.add(c.from_user.id)
        await c.message.answer('🗓 Введите дату: <b>15.08.2026</b>, <b>15.08</b> или <b>2026-08-15</b>.',parse_mode='HTML'); return await c.answer()
    selected=date.today() if choice=='today' else date.today()-timedelta(days=1)
    await c.answer(); await show_day_expenses(c.message,c.from_user.id,selected.isoformat())
@router.callback_query(F.data.startswith('expensecat:'))
async def change_expense_category(c:CallbackQuery):
    if not allowed(c.from_user.id): return await c.answer('Доступ запрещён.',show_alert=True)
    reset_input_state(c.from_user.id); expense_id=int(c.data.split(':')[1]); categories=await db.categories(c.from_user.id)
    await c.message.answer('📂 Выберите новую категорию для этой траты. Ключевые слова не изменятся:',reply_markup=expense_categories_kb(categories,expense_id)); await c.answer()
@router.callback_query(F.data.startswith('expensecatset:'))
async def set_expense_category(c:CallbackQuery):
    if not allowed(c.from_user.id): return await c.answer('Доступ запрещён.',show_alert=True)
    _,expense_id,category_id=c.data.split(':'); category=await db.category(c.from_user.id,int(category_id))
    if not category or not await db.update_expense_category(c.from_user.id,int(expense_id),int(category_id)): return await c.answer('Не удалось изменить категорию траты.',show_alert=True)
    await c.message.edit_text(f'✅ Трата перемещена в категорию «{category[1]}».\nКлючевые слова не изменены.'); await c.answer()
@router.callback_query(F.data=='catlist')
async def category_list(c:CallbackQuery):
    if not allowed(c.from_user.id): return await c.answer('Доступ запрещён.',show_alert=True)
    reset_input_state(c.from_user.id); rows=await db.categories(c.from_user.id)
    await c.message.edit_text('🏷️ Выберите категорию для управления ключевыми словами или удаления:',reply_markup=categories_manage_kb(rows)); await c.answer()
@router.callback_query(F.data.startswith('catview:'))
async def category_view(c:CallbackQuery):
    if not allowed(c.from_user.id): return await c.answer('Доступ запрещён.',show_alert=True)
    reset_input_state(c.from_user.id); category_id=int(c.data.split(':')[1]); category=await db.category(c.from_user.id,category_id)
    if not category: return await c.answer('Категория не найдена.',show_alert=True)
    keywords=', '.join(x.strip() for x in category[2].split(',') if x.strip()) or 'нет'
    await c.message.edit_text(f'🏷️ {category[1]}\n\nКлючевые слова: {keywords}',reply_markup=category_manage_kb(category_id)); await c.answer()
@router.callback_query(F.data=='catnew')
async def category_new(c:CallbackQuery):
    if not allowed(c.from_user.id): return await c.answer('Доступ запрещён.',show_alert=True)
    reset_input_state(c.from_user.id); awaiting_category.add(c.from_user.id)
    await c.message.answer('➕ Введите название новой категории.'); await c.answer()
@router.callback_query(F.data.startswith('catkwadd:'))
async def category_keyword_add(c:CallbackQuery):
    if not allowed(c.from_user.id): return await c.answer('Доступ запрещён.',show_alert=True)
    category_id=int(c.data.split(':')[1]); category=await db.category(c.from_user.id,category_id)
    if not category: return await c.answer('Категория не найдена.',show_alert=True)
    reset_input_state(c.from_user.id); awaiting_keyword[c.from_user.id]=category_id
    await c.message.answer('➕ Введите одно или несколько ключевых слов через запятую.\nНапример: кофе, кофейня, латте'); await c.answer()
@router.callback_query(F.data.startswith('catkwlist:'))
async def category_keyword_list(c:CallbackQuery):
    if not allowed(c.from_user.id): return await c.answer('Доступ запрещён.',show_alert=True)
    reset_input_state(c.from_user.id); category_id=int(c.data.split(':')[1]); category=await db.category(c.from_user.id,category_id)
    if not category: return await c.answer('Категория не найдена.',show_alert=True)
    keywords=[x.strip() for x in category[2].split(',') if x.strip()]
    text=f'✏️ Выберите ключевое слово категории «{category[1]}»:' if keywords else f'У категории «{category[1]}» пока нет ключевых слов.'
    await c.message.edit_text(text,reply_markup=category_keywords_kb(category_id,keywords)); await c.answer()
@router.callback_query(F.data.startswith('catkwedit:'))
@router.callback_query(F.data.startswith('catkwdel:'))
async def category_keyword_edit(c:CallbackQuery):
    if not allowed(c.from_user.id): return await c.answer('Доступ запрещён.',show_alert=True)
    reset_input_state(c.from_user.id); _,category_id,index=c.data.split(':'); category_id=int(category_id); index=int(index)
    category=await db.category(c.from_user.id,category_id); keywords=[x.strip() for x in category[2].split(',') if x.strip()] if category else []
    if index < 0 or index >= len(keywords): return await c.answer('Ключевое слово не найдено.',show_alert=True)
    await c.message.edit_text(f'Ключевое слово: «{keywords[index]}»\nКатегория: {category[1]}\n\nЧто изменить?',reply_markup=keyword_manage_kb(category_id,index)); await c.answer()
@router.callback_query(F.data.startswith('catkwrename:'))
async def category_keyword_rename(c:CallbackQuery):
    if not allowed(c.from_user.id): return await c.answer('Доступ запрещён.',show_alert=True)
    _,category_id,index=c.data.split(':'); category_id=int(category_id); index=int(index); category=await db.category(c.from_user.id,category_id)
    keywords=[x.strip() for x in category[2].split(',') if x.strip()] if category else []
    if index < 0 or index >= len(keywords): return await c.answer('Ключевое слово не найдено.',show_alert=True)
    reset_input_state(c.from_user.id); awaiting_keyword_rename[c.from_user.id]=(category_id,index)
    await c.message.answer(f'✏️ Введите новый текст для ключевого слова «{keywords[index]}».\nОн также заменится в описаниях прошлых трат.'); await c.answer()
@router.callback_query(F.data.startswith('catkwmove:'))
async def category_keyword_move(c:CallbackQuery):
    if not allowed(c.from_user.id): return await c.answer('Доступ запрещён.',show_alert=True)
    reset_input_state(c.from_user.id); _,category_id,index=c.data.split(':'); category_id=int(category_id); index=int(index); category=await db.category(c.from_user.id,category_id)
    keywords=[x.strip() for x in category[2].split(',') if x.strip()] if category else []
    if index < 0 or index >= len(keywords): return await c.answer('Ключевое слово не найдено.',show_alert=True)
    categories=await db.categories(c.from_user.id)
    await c.message.edit_text(f'📂 Куда переместить ключевое слово «{keywords[index]}»?\nВсе прошлые траты с ним тоже сменят категорию.',reply_markup=keyword_move_kb(category_id,index,categories)); await c.answer()
@router.callback_query(F.data.startswith('catkwmoveto:'))
async def category_keyword_move_to(c:CallbackQuery):
    if not allowed(c.from_user.id): return await c.answer('Доступ запрещён.',show_alert=True)
    reset_input_state(c.from_user.id); _,source_id,index,target_id=c.data.split(':'); result=await db.move_category_keyword(c.from_user.id,int(source_id),int(index),int(target_id))
    if not result: return await c.answer('Не удалось переместить ключевое слово.',show_alert=True)
    keyword,target_name,moved=result
    await c.message.edit_text(f'✅ «{keyword}» перемещено в категорию «{target_name}».\nПеремещено прошлых трат: {moved}.',reply_markup=category_manage_kb(int(target_id))); await c.answer()
@router.callback_query(F.data.startswith('catkwdelete:'))
async def category_keyword_delete_confirm(c:CallbackQuery):
    if not allowed(c.from_user.id): return await c.answer('Доступ запрещён.',show_alert=True)
    reset_input_state(c.from_user.id); _,category_id,index=c.data.split(':'); category_id=int(category_id); index=int(index); category=await db.category(c.from_user.id,category_id)
    keywords=[x.strip() for x in category[2].split(',') if x.strip()] if category else []
    if index < 0 or index >= len(keywords): return await c.answer('Ключевое слово не найдено.',show_alert=True)
    await c.message.edit_text(f'Удалить ключевое слово «{keywords[index]}»?\nПрошлые траты останутся без изменений.',reply_markup=keyword_delete_kb(category_id,index)); await c.answer()
@router.callback_query(F.data.startswith('catkwdeleteyes:'))
async def category_keyword_delete(c:CallbackQuery):
    if not allowed(c.from_user.id): return await c.answer('Доступ запрещён.',show_alert=True)
    reset_input_state(c.from_user.id); _,category_id,index=c.data.split(':'); category_id=int(category_id); removed=await db.remove_category_keyword(c.from_user.id,category_id,int(index))
    if removed is None: return await c.answer('Ключевое слово уже удалено.',show_alert=True)
    await c.message.edit_text(f'✅ Ключевое слово «{removed}» удалено.\nПрошлые траты не изменены.',reply_markup=category_manage_kb(category_id)); await c.answer()
@router.callback_query(F.data.startswith('catdelete:'))
async def category_delete_confirm(c:CallbackQuery):
    if not allowed(c.from_user.id): return await c.answer('Доступ запрещён.',show_alert=True)
    reset_input_state(c.from_user.id); category_id=int(c.data.split(':')[1]); category=await db.category(c.from_user.id,category_id)
    if not category: return await c.answer('Категория не найдена.',show_alert=True)
    await c.message.edit_text(f'Удалить категорию «{category[1]}»?\n\nСуществующие траты сохранятся и попадут в «Прочее».',reply_markup=category_delete_kb(category_id)); await c.answer()
@router.callback_query(F.data.startswith('catdeleteyes:'))
async def category_delete(c:CallbackQuery):
    if not allowed(c.from_user.id): return await c.answer('Доступ запрещён.',show_alert=True)
    reset_input_state(c.from_user.id); category_id=int(c.data.split(':')[1]); category=await db.category(c.from_user.id,category_id)
    if not category or not await db.delete_category(c.from_user.id,category_id): return await c.answer('Категория уже удалена.',show_alert=True)
    await c.message.edit_text(f'✅ Категория «{category[1]}» удалена.',reply_markup=InlineKeyboardMarkup(inline_keyboard=[[InlineKeyboardButton(text='⬅️ К категориям',callback_data='catlist')]])); await c.answer()
@router.callback_query(F.data.startswith('choose:'))
async def choose(c:CallbackQuery):
    if not allowed(c.from_user.id): return await c.answer('Доступ запрещён.',show_alert=True)
    if c.from_user.id not in pending: return await c.answer('Этот выбор уже неактуален. Отправьте трату заново.',show_alert=True)
    info=pending.pop(c.from_user.id); cat=await db.category(c.from_user.id,int(c.data.split(':')[1])); expense_id=await db.add_expense(c.from_user.id,cat[0],*info[:4]); await db.learn_keyword(c.from_user.id,cat[0],info[1])
    await c.message.edit_text(f'✅ {info[1]} — {info[0]:g} {info[3]}\nКатегория: {cat[1]}\nДата: {info[2]}',reply_markup=saved_expense_kb(expense_id)); await c.answer()
@router.callback_query(F.data=='newcat')
async def newcat_callback(c:CallbackQuery):
    if not allowed(c.from_user.id): return await c.answer('Доступ запрещён.',show_alert=True)
    reset_input_state(c.from_user.id,keep_pending=True); awaiting_category.add(c.from_user.id); await c.message.answer('Введите название новой категории.'); await c.answer()
@router.message(F.text.in_({'💸 Добавить трату','📊 Отчёт','✏️ Редактировать траты','🏷️ Категории','📅 За день','⚙️ Настройки','➕ Категория','💱 Валюта','🧭 Помощь'}))
async def menu_buttons(m: Message):
    if not await guard(m): return
    reset_input_state(m.from_user.id); action=m.text
    if action=='💸 Добавить трату':
        await m.answer('💸 Пришлите трату: <b>Кофе 250</b>, <b>200 хлеб</b> или <b>Такси 1500 #Транспорт</b>.',parse_mode='HTML')
    elif action=='📊 Отчёт':
        await m.answer('📊 Выберите период отчёта:',reply_markup=report_period_kb())
    elif action=='✏️ Редактировать траты':
        await m.answer('✏️ За какой день показать траты?',reply_markup=edit_expenses_period_kb())
    elif action=='🏷️ Категории':
        await cats(m)
    elif action=='📅 За день':
        await m.answer('📊 Выберите период отчёта:',reply_markup=report_period_kb())
    elif action=='⚙️ Настройки':
        await settings(m)
    elif action=='➕ Категория':
        awaiting_category.add(m.from_user.id); await m.answer('➕ Введите название новой категории.')
    elif action=='💱 Валюта':
        await m.answer('💱 Выберите валюту или введите свою:',reply_markup=currency_kb())
    else:
        await help_(m)
@router.message(F.text)
async def text(m:Message):
    if not await guard(m): return
    uid=m.from_user.id
    if uid in awaiting_expense_day:
        try: selected=parse_day(m.text)
        except Exception: return await m.answer('Введите дату в формате 15.08, 15.08.2026 или 2026-08-15.')
        awaiting_expense_day.remove(uid)
        await show_day_expenses(m,uid,selected)
        return
    if uid in awaiting_report_period:
        try: start,end=parse_period(m.text)
        except ValueError: return await m.answer('Введите период в формате: <b>01.08.2026 - 15.08.2026</b>',parse_mode='HTML')
        awaiting_report_period.remove(uid)
        await ask_report_category(m,uid,start,end)
        return
    if uid in awaiting_currency:
        selected=m.text.strip()
        if not selected or len(selected)>12: return await m.answer('Введите короткое обозначение валюты — например: €, EUR или сом.')
        awaiting_currency.remove(uid)
        await db.set_currency(uid,selected)
        await m.answer(f'✅ Валюта изменена на {selected}.')
        return
    if uid in awaiting_amount:
        expense_id=awaiting_amount[uid]
        try: amount=float(m.text.strip().replace(',','.')); assert amount>0
        except Exception: return await m.answer('Введите положительную сумму, например: 350.50')
        awaiting_amount.pop(uid)
        if await db.update_expense_amount(uid,expense_id,amount): await m.answer(f'✅ Сумма изменена на {amount:g}.')
        else: await m.answer('Не нашёл эту трату.')
        return
    if uid in awaiting_date:
        expense_id=awaiting_date[uid]
        try: expense_date=parse_day(m.text)
        except Exception: return await m.answer('Введите дату в формате 03.08, 03.08.2026 или 2026-08-03.')
        awaiting_date.pop(uid)
        if await db.update_expense_date(uid,expense_id,expense_date): await m.answer(f'✅ Трата перенесена на {expense_date}.')
        else: await m.answer('Не нашёл эту трату.')
        return
    if uid in awaiting_keyword_rename:
        category_id,index=awaiting_keyword_rename[uid]; replacement=m.text.strip()
        if not replacement or len(replacement)>50 or ',' in replacement: return await m.answer('Введите одно ключевое слово или фразу без запятой, не длиннее 50 символов.')
        awaiting_keyword_rename.pop(uid)
        result=await db.rename_category_keyword(uid,category_id,index,replacement)
        if not result: return await m.answer('Не удалось изменить ключевое слово. Возможно, такое слово уже есть в категории.',reply_markup=category_manage_kb(category_id))
        old,new,changed=result
        await m.answer(f'✅ «{old}» изменено на «{new}».\nОбновлено прошлых трат: {changed}.',reply_markup=category_manage_kb(category_id))
        return
    if uid in awaiting_keyword:
        category_id=awaiting_keyword[uid]
        keywords=[x.strip() for x in m.text.split(',') if x.strip()]
        if not keywords or any(len(x)>50 for x in keywords): return await m.answer('Введите ключевые слова через запятую, каждое не длиннее 50 символов.')
        awaiting_keyword.pop(uid)
        added=await db.add_category_keywords(uid,category_id,m.text)
        category=await db.category(uid,category_id)
        if not category: return await m.answer('Категория не найдена.')
        if added: await m.answer(f'✅ Добавлены ключевые слова: {", ".join(added)}',reply_markup=category_manage_kb(category_id))
        else: await m.answer('Все эти ключевые слова уже есть у категории.',reply_markup=category_manage_kb(category_id))
        return
    if uid in awaiting_category:
        awaiting_category.remove(uid)
        try:
            category_id=await db.add_category(uid,m.text.strip())
            if uid in pending:
                info=pending.pop(uid); expense_id=await db.add_expense(uid,category_id,*info[:4]); await db.learn_keyword(uid,category_id,info[1])
                await m.answer(f'✅ {info[1]} — {info[0]:g} {info[3]}\nНовая категория: {m.text.strip()}\nДата: {info[2]}',reply_markup=saved_expense_kb(expense_id))
            else: await m.answer(f'Категория «{m.text.strip()}» добавлена.')
        except Exception: await m.answer('Такая категория уже есть.')
        return
    try: description,amount,d,currency,manual=parse_expense(m.text)
    except ValueError as e:return await m.answer(str(e))
    d=day_override.pop(uid,d)
    currency=currency or await db.currency(uid); rows=await db.categories(uid)
    cat=next((x for x in rows if x[1].lower()==manual.lower()),None) if manual else await db.find_category(uid,description)
    if cat:
        expense_id=await db.add_expense(uid,cat[0],amount,description,d,currency); return await m.answer(f'✅ {description} — {amount:g} {currency}\nКатегория: {cat[1]}\nДата: {d}',reply_markup=saved_expense_kb(expense_id))
    pending[uid]=(amount,description,d,currency); await m.answer('Не удалось определить категорию. Выберите:',reply_markup=categories_kb(rows))
async def main():
    if not TOKEN or TOKEN=='replace_me': raise RuntimeError('Fill BOT_TOKEN in .env')
    if not ALLOWED: raise RuntimeError('Set ALLOWED_USER_IDS in .env')
    if not ENCRYPTION_KEY: raise RuntimeError('Set ENCRYPTION_KEY in .env or hosting environment')
    await db.init(); bot=Bot(TOKEN); await bot.set_my_commands(MENU_COMMANDS)
    for user_id in ALLOWED:
        try:
            await bot.send_message(user_id, '✅ Бот запущен. Главное меню — внизу 👇', reply_markup=MAIN_MENU)
        except TelegramAPIError as exc:
            logging.warning('Could not notify allowed user %s: %s', user_id, exc)
    dp=Dispatcher(); dp.include_router(router); reminder_task=asyncio.create_task(reminder_loop(bot))
    try: await dp.start_polling(bot)
    finally: reminder_task.cancel()
if __name__=='__main__': logging.basicConfig(level=logging.INFO); asyncio.run(main())
